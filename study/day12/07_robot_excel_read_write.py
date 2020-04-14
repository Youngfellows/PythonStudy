# coding=utf-8
# 读写2003 excel
import xlrd
import xlwt

# 读写2007 excel
import openpyxl
import json
import os

# 修改（追加写入）
from xlutils.copy import copy


class Excel(object):
    '''Excel表格读写'''

    def __init__(self):
        object.__init__(self)
        self.car_table_path = "./json/car_info.json"
        self.phone_table_path = "./json/phone_info.json"
        self.driver_table_path = "./json/driver_info.json"
        self.passenger_table_path = "./json/passenger_info.json"

    def write_excel_xls(self, path, sheet_name, value):
        '''写入到Excel 2003版本xls格式表格'''
        index = len(value)  # 获取需要写入数据的行数
        workbook = xlwt.Workbook()  # 新建一个工作簿
        sheet = workbook.add_sheet(sheet_name)  # 在工作簿中新建一个表格
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.write(i, j, value[i][j])  # 向表格中写入数据（对应的行和列）
        workbook.save(path)  # 保存工作簿
        print("Excel 2003 xls格式表格写入数据成功！")

    def write_excel_xls_append(self, path, value):
        '''向Excel 2003版本的xls格式表格追加数据'''
        index = len(value)  # 获取需要写入数据的行数
        workbook = xlrd.open_workbook(path)  # 打开工作簿
        sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
        worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
        rows_old = worksheet.nrows  # 获取表格中已存在的数据的行数
        new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
        new_worksheet = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
        for i in range(0, index):
            for j in range(0, len(value[i])):
                new_worksheet.write(i + rows_old, j, value[i][j])  # 追加写入数据，注意是从i+rows_old行开始写入
        new_workbook.save(path)  # 保存工作簿
        print("xls格式表格【追加】写入数据成功！")

    def read_excel_xls(self, path):
        '''读取Excel2003版本的xls格式表格'''
        workbook = xlrd.open_workbook(path)  # 打开工作簿
        sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
        worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
        for i in range(0, worksheet.nrows):
            for j in range(0, worksheet.ncols):
                print(worksheet.cell_value(i, j), "\t", end="")  # 逐行逐列读取数据
            print()

    def write_excel_xlsx(self, path, sheet_name, value):
        '''写入到Excel 2007版本xlsx格式表格'''
        index = len(value)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
        workbook.save(path)
        print("Excel 2077 xlsx格式表格写入数据成功！")

    def read_excel_xlsx(self, path, sheet_name):
        '''读取Excel2007版本的xlsx格式表格'''
        workbook = openpyxl.load_workbook(path)
        # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
        sheet = workbook[sheet_name]
        for row in sheet.rows:
            for cell in row:
                print(cell.value, "\t", end="")
            print()

    def parse_robot_excel_xlsx(self, path):
        '''解析Robot数据规则表'''
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)  # 获取工作布
        sheet_names = workbook.sheetnames  # 获取sheet表名
        print(sheet_names)
        for sheet_name in sheet_names:
            if sheet_name == "车辆":
                print("*" * 60)
                print("car shell ... ")
                self.parse_car_sheet(sheet_name, workbook)
            elif sheet_name == "司机":
                print("*" * 100)
                print("driver shell ... ")
                self.parse_driver_sheet(sheet_name, workbook)
            elif sheet_name == "乘客":
                print("*" * 100)
                print("passenger shell ... ")
                self.parse_passenger_sheet(sheet_name, workbook)
            elif sheet_name == "Sheet1":
                '''imei 表'''
                print("*" * 100)
                print("imei shell ... ")
                self.parse_imei_sheet(sheet_name, workbook)

    def parse_passenger_sheet(self, sheet_name, workbook):
        '''解析passenger sheet'''
        passenger_list = []  # 列表
        passenger_sn = "serialNumber"  # 编号
        name = "name"  # 姓名
        nickName = "nickName"  # 昵称
        birthday = "birthday"  # 生日
        phone_number = "phoneNumber"  # 电话号码
        sheet = workbook[sheet_name]  # 获取到具体表
        for i, row in enumerate(sheet.rows):
            passenger = {}  # 字典对象
            for j, cell in enumerate(row):
                cell_value = cell.value
                print(r"({0} {1}) {2}".format(i, j, cell_value), "\t", end="")
                if i != 0:
                    if j == 0:
                        passenger[passenger_sn] = cell_value
                    elif j == 1:
                        passenger[name] = cell_value
                    elif j == 2:
                        passenger[nickName] = cell_value
                    elif j == 3:
                        passenger[birthday] = str(cell_value)
                    elif j == 4:
                        passenger[phone_number] = cell_value
            # 首行不加入列表
            if passenger:
                passenger_list.append(passenger)
            print()
        # dumps()默认中文为ascii编码格式，ensure_ascii默认为Ture
        # 禁用ascii编码格式，返回的Unicode字符串，方便使用
        json_str = json.dumps(passenger_list, ensure_ascii=False)
        print("Passenger列表: json_str = {}".format(json_str))
        # 保存
        self.writer2json(self.passenger_table_path, passenger_list)

    def parse_driver_sheet(self, sheet_name, workbook):
        '''解析driver sheet'''
        driver_list = []  # 列表
        driver_sn = "serialNumber"  # 编号
        name = "name"  # 姓名
        idCard = "idCard"  # 身份证号
        birthday = "birthday"  # 生日
        driving_licence = "drivingLicence"  # 驾照号
        phone_number = "phoneNumber"  # 电话号码
        contact_address = "contactAddress"  # 联络地址
        address = "address"  # 户籍地址
        sheet = workbook[sheet_name]  # 获取到具体表
        for i, row in enumerate(sheet.rows):
            driver = {}  # 字典对象
            for j, cell in enumerate(row):
                cell_value = cell.value
                print(r"({0} {1}) {2}".format(i, j, cell_value), "\t", end="")
                if i != 0:
                    if j == 0:
                        driver[driver_sn] = cell_value
                    elif j == 1:
                        driver[name] = cell_value
                    elif j == 2:
                        driver[idCard] = cell_value
                    elif j == 3:
                        driver[birthday] = str(cell_value)
                    elif j == 4:
                        driver[driving_licence] = cell_value
                    elif j == 8:
                        driver[phone_number] = cell_value
                    elif j == 9:
                        driver[contact_address] = cell_value
                    elif j == 10:
                        driver[address] = cell_value
            # 首行不加入列表
            if driver:
                driver_list.append(driver)
            print()
        # dumps()默认中文为ascii编码格式，ensure_ascii默认为Ture
        # 禁用ascii编码格式，返回的Unicode字符串，方便使用
        json_str = json.dumps(driver_list, ensure_ascii=False)
        print("Driver列表: json_str = {}".format(json_str))
        # 保存
        self.writer2json(self.driver_table_path, driver_list)

    def parse_imei_sheet(self, sheet_name, workbook):
        '''解析phone sheet'''
        phone_list = []  # 列表
        phone_brand = "brand"  # 品牌
        phone_androidId = "androidId"  # Android ID
        phone_imei = "imei"  # imei
        phone_sn = "serialNumber"  # 序列号
        phone_car = "mac"  # MAC地址
        sheet = workbook[sheet_name]  # 获取到具体表
        for i, row in enumerate(sheet.rows):
            phone = {}  # 字典对象
            for j, cell in enumerate(row):
                cell_value = cell.value
                print(r"({0} {1}) {2}".format(i, j, cell_value), "\t", end="")
                if i != 0:
                    if j == 0:
                        phone[phone_brand] = cell_value
                    elif j == 1:
                        phone[phone_androidId] = cell_value
                    elif j == 2:
                        phone[phone_imei] = cell_value
                    elif j == 3:
                        phone[phone_sn] = cell_value
                    elif j == 4:
                        phone[phone_car] = cell_value
            # 首行不加入列表
            if phone:
                phone_list.append(phone)
            print()
        # dumps()默认中文为ascii编码格式，ensure_ascii默认为Ture
        # 禁用ascii编码格式，返回的Unicode字符串，方便使用
        json_str = json.dumps(phone_list, ensure_ascii=False)
        print("Phone列表: json_str = {}".format(json_str))
        # 保存
        self.writer2json(self.phone_table_path, phone_list)

    def parse_car_sheet(self, sheet_name, workbook):
        '''解析car sheet'''
        car_list = []  # 列表
        car_number = "carNumber"  # 车辆编号
        car_vin = "vin"  # 车架号
        car_plateNumber = "plateNumber"  # 车牌号
        car_brand = "brand"  # 品牌
        car_colour = "colour"  # 车颜色
        sheet = workbook[sheet_name]  # 获取到具体表
        for i, row in enumerate(sheet.rows):
            car = {}  # 字典对象
            for j, cell in enumerate(row):
                cell_value = cell.value
                print(r"({0} {1}) {2}".format(i, j, cell_value), "\t", end="")
                if i != 0:
                    if j == 0:
                        car[car_number] = cell_value
                    elif j == 1:
                        car[car_vin] = cell_value
                    elif j == 4:
                        car[car_plateNumber] = cell_value
                    elif j == 5:
                        car[car_brand] = cell_value
                    elif j == 8:
                        car[car_colour] = cell_value
            # 首行不加入列表
            if car:
                car_list.append(car)
            print()
        # dumps()默认中文为ascii编码格式，ensure_ascii默认为Ture
        # 禁用ascii编码格式，返回的Unicode字符串，方便使用
        json_str = json.dumps(car_list, ensure_ascii=False)
        print("车辆列表: json_str = {}".format(json_str))
        # 保存
        self.writer2json(self.car_table_path, car_list)

    def json2dict(self, json_data):
        """将json字符串转化为python的字典对象"""
        return json.loads(json_data)

    def read2json(self, file_name):
        """读取json文件,并转换为字典/列表"""
        with open(file_name, "r", encoding="utf-8") as fp:
            dict = json.load(fp)
        print(dict)
        return dict

    def writer2json(self, file_name, dict):
        """将字典对象保存为json字符串"""
        # 删除旧文件
        if file_name in os.listdir():
            os.remove(file_name)

        # dumps()默认中文为ascii编码格式，ensure_ascii默认为Ture
        # 禁用ascii编码格式，返回的Unicode字符串，方便使用
        json_str = json.dumps(dict, ensure_ascii=False)
        with open(file_name, "wb") as fp:
            fp.write(json_str.encode('utf-8'))


def parse_car_sheet(self):
    '''解析car sheet'''
    pass


def parse_driver_sheet(self):
    '''解析driver sheet'''
    pass


def parse_passenger_sheet(self):
    '''解析passenger sheet'''
    pass


def parse_imei_sheet(self):
    '''解析imei sheet'''
    pass


if __name__ == "__main__":
    excel = Excel()
    # Excel 2003 xls格式表格
    book_name_xls = 'Excel 2003 xls格式测试工作簿.xls'  # 文件名
    sheet_name_xls = 'xls格式测试表'  # 表名

    value_title = [["姓名", "性别", "年龄", "城市", "职业"]]

    value1 = [["张三", "男", "19", "杭州", "研发工程师"],
              ["李四", "男", "22", "北京", "医生"],
              ["王五", "女", "33", "珠海", "出租车司机"]]

    value2 = [["Tom", "男", "21", "西安", "测试工程师"],
              ["Jones", "女", "34", "上海", "产品经理"],
              ["Cat", "女", "56", "上海", "教师"]]
    file_2003_path = "./excel/" + book_name_xls
    excel.write_excel_xls(file_2003_path, sheet_name_xls, value1)
    excel.write_excel_xls_append(file_2003_path, value2)
    excel.read_excel_xls(file_2003_path)

    # Excel 2007 xls格式表格
    book_name_xlsx = 'Excel 2007 xlsx格式测试工作簿.xlsx'
    sheet_name_xlsx = 'xlsx格式测试表'
    value3 = [["姓名", "性别", "年龄", "城市", "职业"],
              ["111", "女", "66", "石家庄", "运维工程师"],
              ["222", "男", "55", "南京", "饭店老板"],
              ["333", "女", "27", "苏州", "保安"], ]
    file_2007_path = "./excel/" + book_name_xlsx
    excel.write_excel_xlsx(file_2007_path, sheet_name_xlsx, value3)
    excel.read_excel_xlsx(file_2007_path, sheet_name_xlsx)

    # 操作Robot表,imei
    book_name_robot_data = "数据及规则.xlsx"
    book_name_robot_imei = "RobotIMEI.xlsx"
    file_robot_data_path = "./excel/" + book_name_robot_data
    file_robot_imei_path = "./excel/" + book_name_robot_imei
    excel.parse_robot_excel_xlsx(file_robot_data_path)
    excel.parse_robot_excel_xlsx(file_robot_imei_path)
