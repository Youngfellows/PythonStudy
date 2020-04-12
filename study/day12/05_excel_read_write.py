# coding=utf-8
# 读写2003 excel
import xlrd
import xlwt

# 读写2007 excel
import openpyxl


class Excel(object):
    '''Excel表格读写'''

    def __init__(self):
        object.__init__(self)

    def write03Excel(self, path):
        '''写入到2003版本的Excel'''
        wb = xlwt.Workbook()
        sheet = wb.add_sheet('2003测试表')
        value = [["名称", "价格", "出版社", "语言"],
                 ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
                 ["暗时间", "32.4", "人民邮电出版社", "中文"],
                 ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]

        for i in range(0, 4):
            for j in range(0, len(value[i])):
                sheet.write(i, j, value[i][j])
        wb.save(path)
        print("Excel 2003 写入数据成功！")

    def read03Excel(self, path):
        '''读取Excel2003版本的Excel表格'''
        workbook = xlrd.open_workbook(path)
        sheets = workbook.sheet_names()
        worksheet = workbook.sheet_by_name(sheets[0])
        for i in range(0, worksheet.nrows):
            # 每一行
            row = worksheet.row(i)
            for j in range(0, worksheet.ncols):
                # 每一行的每一列数据
                value = worksheet.cell_value(i, j)
                print("{}\t".format(value), end="")
            print()

    def write07Excel(self, path):
        '''写入到2007版本的Excel'''
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = '2007测试表'  # 表名
        value = [["名称", "价格", "出版社", "语言"],
                 ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
                 ["暗时间", "32.4", "人民邮电出版社", "中文"],
                 ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
        # print("len:{}".format(len(value)))

        for i in range(0, len(value)):
            for j in range(0, len(value[i])):
                # 写入表格
                sheet.cell(i + 1, j + 1, value=str(value[i][j]))
        wb.save(path)
        print("Excel 2007 写入数据成功！")

    def read07Excel(self, path):
        '''读取Excel2007版本的Excel表格'''
        wb = openpyxl.load_workbook(path)
        # sheet = wb.get_sheet_by_name("2007测试表")  # 获取到表
        sheet = wb["2007测试表"]  # 获取到表
        # 遍历表
        for row in sheet.rows:
            for cell in row:
                value = cell.value
                print("{}\t".format(value), end="")
            print()


if __name__ == "__main__":
    excel = Excel()

    file_2003_path = "./excel/2003.xls"
    file_2007_path = "./excel/2007.xlsx"
    excel.write03Excel(file_2003_path)
    excel.read03Excel(file_2003_path)

    excel.write07Excel(file_2007_path)
    excel.read07Excel(file_2007_path)
