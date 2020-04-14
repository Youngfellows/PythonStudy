# coding=utf-8
# 读写2003 excel
import xlrd
import xlwt

# 读写2007 excel
import openpyxl

# 修改（追加写入）
from xlutils.copy import copy


class Excel(object):
    '''Excel表格读写'''

    def __init__(self):
        object.__init__(self)

    def write_excel_xls(self, path, sheet_name, value):
        '''写入到Excel 2003版本xls格式表格'''
        index = len(value)  # 获取需要写入数据的行数
        workbook = xlwt.Workbook()  # 新建一个工作簿
        sheet = workbook.add_sheet(sheet_name)  # 在工作簿中新建一个表格
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.write(i, j, value[i][j])  # 向表格中写入数据（对应的行和列）
        workbook.save(path)  # 保存工作簿
        print("Excel 2003 xls格式表格写入数据成功！")

    def write_excel_xls_append(self, path, value):
        '''向Excel 2003版本的xls格式表格追加数据'''
        index = len(value)  # 获取需要写入数据的行数
        workbook = xlrd.open_workbook(path)  # 打开工作簿
        sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
        worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
        rows_old = worksheet.nrows  # 获取表格中已存在的数据的行数
        new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
        new_worksheet = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
        for i in range(0, index):
            for j in range(0, len(value[i])):
                new_worksheet.write(i + rows_old, j, value[i][j])  # 追加写入数据，注意是从i+rows_old行开始写入
        new_workbook.save(path)  # 保存工作簿
        print("xls格式表格【追加】写入数据成功！")

    def read_excel_xls(self, path):
        '''读取Excel2003版本的xls格式表格'''
        workbook = xlrd.open_workbook(path)  # 打开工作簿
        sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
        worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
        for i in range(0, worksheet.nrows):
            for j in range(0, worksheet.ncols):
                print(worksheet.cell_value(i, j), "\t", end="")  # 逐行逐列读取数据
            print()

    def write_excel_xlsx(self, path, sheet_name, value):
        '''写入到Excel 2007版本xlsx格式表格'''
        index = len(value)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
        workbook.save(path)
        print("Excel 2077 xlsx格式表格写入数据成功！")

    def read_excel_xlsx(self, path, sheet_name):
        '''读取Excel2007版本的xlsx格式表格'''
        workbook = openpyxl.load_workbook(path)
        # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
        sheet = workbook[sheet_name]
        for row in sheet.rows:
            for cell in row:
                print(cell.value, "\t", end="")
            print()


if __name__ == "__main__":
    excel = Excel()
    # Excel 2003 xls格式表格
    book_name_xls = 'Excel 2003 xls格式测试工作簿.xls'  # 文件名
    sheet_name_xls = 'xls格式测试表'  # 表名

    value_title = [["姓名", "性别", "年龄", "城市", "职业"]]

    value1 = [["张三", "男", "19", "杭州", "研发工程师"],
              ["李四", "男", "22", "北京", "医生"],
              ["王五", "女", "33", "珠海", "出租车司机"]]

    value2 = [["Tom", "男", "21", "西安", "测试工程师"],
              ["Jones", "女", "34", "上海", "产品经理"],
              ["Cat", "女", "56", "上海", "教师"]]
    file_2003_path = "./excel/" + book_name_xls
    excel.write_excel_xls(file_2003_path, sheet_name_xls, value1)
    excel.write_excel_xls_append(file_2003_path, value2)
    excel.read_excel_xls(file_2003_path)

    # Excel 2007 xls格式表格
    book_name_xlsx = 'Excel 2007 xlsx格式测试工作簿.xlsx'
    sheet_name_xlsx = 'xlsx格式测试表'
    value3 = [["姓名", "性别", "年龄", "城市", "职业"],
              ["111", "女", "66", "石家庄", "运维工程师"],
              ["222", "男", "55", "南京", "饭店老板"],
              ["333", "女", "27", "苏州", "保安"], ]
    file_2007_path = "./excel/" + book_name_xlsx
    excel.write_excel_xlsx(file_2007_path, sheet_name_xlsx, value3)
    excel.read_excel_xlsx(file_2007_path, sheet_name_xlsx)
